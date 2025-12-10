"""
Excel file parser for CNJ process numbers
"""
import logging
from typing import List, Set
import openpyxl
from io import BytesIO
import re

logger = logging.getLogger(__name__)


def is_valid_cnj(cnj: str) -> bool:
    """
    Validate CNJ format: NNNNNNN-DD.AAAA.J.TR.OOOO

    Args:
        cnj: CNJ process number

    Returns:
        True if valid CNJ format
    """
    # CNJ pattern: 7 digits - 2 digits . 4 digits . 1 digit . 2 digits . 4 digits
    pattern = r'^\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4}$'
    return bool(re.match(pattern, cnj.strip()))


def clean_cnj(cnj: str) -> str:
    """
    Clean and format CNJ number

    Args:
        cnj: Raw CNJ string

    Returns:
        Cleaned CNJ string
    """
    # Remove extra whitespace
    cnj = cnj.strip()

    # Remove any non-digit/non-dash/non-dot characters
    cnj = re.sub(r'[^\d\-.]', '', cnj)

    return cnj


async def parse_excel_cnjs(file_content: bytes) -> List[str]:
    """
    Parse Excel file and extract CNJ process numbers
    Looks for a column named "CNJ" in the first sheet header

    Args:
        file_content: Excel file content in bytes

    Returns:
        List of unique valid CNJ numbers

    Raises:
        ValueError: If file cannot be parsed or no valid CNJs found
    """
    try:
        # Load workbook from bytes
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

        # Get first sheet (always use first sheet)
        sheet = workbook.worksheets[0]

        cnjs: Set[str] = set()
        invalid_cnjs: List[str] = []

        # Find CNJ column (look for header "CNJ" in first row)
        cnj_column = None
        header_row = 1

        for col_idx, cell in enumerate(sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().upper() == "CNJ":
                cnj_column = col_idx
                logger.info(f"Found CNJ column at index {col_idx}")
                break

        if cnj_column:
            # Parse CNJs from the specific column
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=cnj_column)

                if cell.value:
                    cell_value = str(cell.value).strip()

                    # Skip empty cells
                    if not cell_value:
                        continue

                    # Clean CNJ
                    cleaned_cnj = clean_cnj(cell_value)

                    # Validate CNJ
                    if is_valid_cnj(cleaned_cnj):
                        cnjs.add(cleaned_cnj)
                    else:
                        # Check if it looks like a CNJ
                        if re.search(r'\d{5,}', cell_value):
                            invalid_cnjs.append(cell_value)
        else:
            # Fallback: search all cells if no CNJ column found
            logger.warning("CNJ column not found in header, searching all cells...")

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip()

                        # Skip empty cells or headers
                        if not cell_value or cell_value.upper() in ['CNJ', 'PROCESSO', 'NÚMERO']:
                            continue

                        # Clean CNJ
                        cleaned_cnj = clean_cnj(cell_value)

                        # Validate CNJ
                        if is_valid_cnj(cleaned_cnj):
                            cnjs.add(cleaned_cnj)
                        else:
                            # Check if it looks like a CNJ
                            if re.search(r'\d{5,}', cell_value):
                                invalid_cnjs.append(cell_value)

        # Close workbook
        workbook.close()

        if not cnjs:
            error_msg = "Nenhum número CNJ válido encontrado no arquivo Excel"
            if invalid_cnjs:
                error_msg += f". Encontrados {len(invalid_cnjs)} CNJs inválidos. Exemplos: {invalid_cnjs[:3]}"
            else:
                error_msg += ". Certifique-se de que a planilha tem uma coluna 'CNJ' com números de processo válidos."
            raise ValueError(error_msg)

        logger.info(f"Parsed {len(cnjs)} valid CNJs from Excel file")

        if invalid_cnjs:
            logger.warning(f"Found {len(invalid_cnjs)} invalid CNJ entries")

        return sorted(list(cnjs))

    except openpyxl.utils.exceptions.InvalidFileException:
        raise ValueError("Formato de arquivo Excel inválido. Por favor, faça upload de um arquivo .xlsx válido")
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise ValueError(f"Erro ao processar arquivo Excel: {str(e)}")
